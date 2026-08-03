"""Tests for the two-workbook Excel input loader."""

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from pgl_sorting_engine import (
    LocationName,
    SpreadsheetValidationError,
    load_sorting_workbooks,
)


def create_configuration_workbook(path: Path) -> None:
    """Create a minimal valid configuration workbook."""
    workbook = Workbook()

    pathologists = workbook.active
    pathologists.title = "Pathologists"
    pathologists.append(
        [
            "pathologist_id",
            "display_name",
            "subspecialties",
        ]
    )
    pathologists.append(
        [
            "JS",
            "Dr. Smith",
            "GI; LIVER",
        ]
    )

    case_types = workbook.create_sheet("CaseTypes")
    case_types.append(
        [
            "case_type",
            "subspecialty",
            "requirement",
        ]
    )
    case_types.append(
        [
            "GI",
            "GI",
            "required",
        ]
    )

    prefixes = workbook.create_sheet("Prefixes")
    prefixes.append(
        [
            "prefix",
            "allowed_locations",
            "required_location",
            "preferred_locations",
        ]
    )
    prefixes.append(
        [
            "AB",
            "OLOL",
            None,
            None,
        ]
    )

    hospitals = workbook.create_sheet("Hospitals")
    hospitals.append(
        [
            "hospital",
            "allowed_locations",
            "required_location",
        ]
    )
    hospitals.append(
        [
            "Hospital A",
            "OLOL",
            None,
        ]
    )
    assignment_settings = workbook.create_sheet(
        "AssignmentSettings"
    )
    assignment_settings.append(
        [
            "met_weight_per_pathologist",
            "wh_starting_weight",
        ]
    )
    assignment_settings.append([125, 375])

    workbook.save(path)


def create_daily_workbook(path: Path) -> None:
    """Create a minimal valid daily workbook."""
    workbook = Workbook()

    accessions = workbook.active
    accessions.title = "Accessions"
    accessions.append(
        [
            "accession_number",
            "prefix",
            "case_type",
            "hospital",
            "weight",
        ]
    )
    accessions.append(
        [
            "S26-1001",
            "AB",
            "GI",
            "Hospital A",
            2.5,
        ]
    )

    staffing = workbook.create_sheet("Staffing")
    staffing.append(
        [
            "location",
            "pathologist_id",
        ]
    )
    staffing.append(
        [
            "OLOL",
            "JS",
        ]
    )

    workbook.save(path)


def test_valid_workbooks_build_and_run_engine(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    data = load_sorting_workbooks(
        configuration_path=configuration_path,
        daily_path=daily_path,
    )

    result = data.build_engine().run(data.accessions)

    assert len(data.configuration.pathologists) == 1
    assert len(data.daily.accessions) == 1
    assert data.configuration.assignment_settings.met_weight_per_pathologist == (
        Decimal("125")
    )
    assert data.configuration.assignment_settings.wh_starting_weight == (
        Decimal("375")
    )
    assert data.build_engine().settings == (
        data.configuration.assignment_settings
    )
    assert result.assigned_accession_count == 1
    assert result.assignments[0].location is LocationName.OLOL


def test_missing_configuration_sheet_is_identified(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    workbook = load_workbook(configuration_path)
    workbook.remove(workbook["CaseTypes"])
    workbook.save(configuration_path)

    with pytest.raises(
        SpreadsheetValidationError,
    ) as exc_info:
        load_sorting_workbooks(
            configuration_path=configuration_path,
            daily_path=daily_path,
        )

    message = str(exc_info.value)

    assert "CONFIGURATION / CaseTypes" in message
    assert "Required worksheet is missing" in message



def test_missing_assignment_settings_sheet_is_identified(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    workbook = load_workbook(configuration_path)
    workbook.remove(workbook["AssignmentSettings"])
    workbook.save(configuration_path)

    with pytest.raises(SpreadsheetValidationError) as exc_info:
        load_sorting_workbooks(
            configuration_path=configuration_path,
            daily_path=daily_path,
        )

    message = str(exc_info.value)
    assert "CONFIGURATION / AssignmentSettings" in message
    assert "Required worksheet is missing" in message


def test_invalid_assignment_setting_identifies_row(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    workbook = load_workbook(configuration_path)
    workbook["AssignmentSettings"]["A2"] = "not-a-number"
    workbook.save(configuration_path)

    with pytest.raises(SpreadsheetValidationError) as exc_info:
        load_sorting_workbooks(
            configuration_path=configuration_path,
            daily_path=daily_path,
        )

    message = str(exc_info.value)
    assert "CONFIGURATION / AssignmentSettings row 2" in message
    assert "not-a-number" in message

def test_missing_daily_sheet_is_identified(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    workbook = load_workbook(daily_path)
    workbook.remove(workbook["Staffing"])
    workbook.save(daily_path)

    with pytest.raises(
        SpreadsheetValidationError,
    ) as exc_info:
        load_sorting_workbooks(
            configuration_path=configuration_path,
            daily_path=daily_path,
        )

    assert "DAILY / Staffing" in str(exc_info.value)


def test_unknown_staffed_pathologist_identifies_daily_row(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    workbook = load_workbook(daily_path)
    workbook["Staffing"]["B2"] = "ZZ"
    workbook.save(daily_path)

    with pytest.raises(
        SpreadsheetValidationError,
    ) as exc_info:
        load_sorting_workbooks(
            configuration_path=configuration_path,
            daily_path=daily_path,
        )

    message = str(exc_info.value)

    assert "DAILY / Staffing row 2" in message
    assert "ZZ" in message


def test_unknown_case_type_identifies_accession_row(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    workbook = load_workbook(daily_path)
    workbook["Accessions"]["C2"] = "ZZ"
    workbook.save(daily_path)

    with pytest.raises(
        SpreadsheetValidationError,
    ) as exc_info:
        load_sorting_workbooks(
            configuration_path=configuration_path,
            daily_path=daily_path,
        )

    message = str(exc_info.value)

    assert "DAILY / Accessions row 2" in message
    assert "case type ZZ" in message


def test_duplicate_accession_identifies_original_row(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    workbook = load_workbook(daily_path)
    workbook["Accessions"].append(
        [
            "S26-1001",
            "AB",
            "GI",
            "Hospital A",
            1,
        ]
    )
    workbook.save(daily_path)

    with pytest.raises(
        SpreadsheetValidationError,
    ) as exc_info:
        load_sorting_workbooks(
            configuration_path=configuration_path,
            daily_path=daily_path,
        )

    message = str(exc_info.value)

    assert "DAILY / Accessions row 3" in message
    assert "first seen on row 2" in message


def test_semicolon_separated_locations_are_supported(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    workbook = load_workbook(configuration_path)
    workbook["Prefixes"]["B2"] = "OLOL; BRG"
    workbook["Hospitals"]["B2"] = "OLOL; BRG"
    workbook.save(configuration_path)

    data = load_sorting_workbooks(
        configuration_path=configuration_path,
        daily_path=daily_path,
    )

    assert data.configuration.prefix_rules[
        0
    ].allowed_locations == frozenset(
        {
            LocationName.OLOL,
            LocationName.BRG,
        }
    )

def test_optional_routing_overrides_are_loaded(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_workbook(configuration_path)
    create_daily_workbook(daily_path)

    workbook = load_workbook(configuration_path)
    overrides = workbook.create_sheet("RoutingOverrides")
    overrides.append(
        [
            "rule_name",
            "hospital",
            "prefix",
            "case_type",
            "routing_mode",
            "destination_location",
            "preferred_locations",
            "required_subspecialty",
            "weight_cap",
        ]
    )
    overrides.append(
        [
            "AB-GI to OLOL up to 40",
            None,
            "AB",
            "GI",
            "preferred_until_weight_cap",
            "OLOL",
            None,
            None,
            40,
        ]
    )
    workbook.save(configuration_path)

    data = load_sorting_workbooks(
        configuration_path=configuration_path,
        daily_path=daily_path,
    )

    assert len(data.configuration.override_rules) == 1
    rule = data.configuration.override_rules[0]
    assert rule.rule_name == "AB-GI to OLOL up to 40"
    assert rule.weight_cap == Decimal("40")
