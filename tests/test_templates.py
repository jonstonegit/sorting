"""Tests for generated Excel workbook templates."""

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from pgl_sorting_engine import (
    create_configuration_template,
    create_daily_template,
    create_sorting_templates,
    load_sorting_workbooks,
)


def test_create_sorting_templates_creates_both_files(
    tmp_path: Path,
) -> None:
    configuration_path, daily_path = create_sorting_templates(
        tmp_path
    )

    assert configuration_path.exists()
    assert daily_path.exists()
    assert configuration_path.name == "sorting_configuration.xlsx"
    assert daily_path.name == "daily_sorting.xlsx"


def test_configuration_template_has_expected_sheets(
    tmp_path: Path,
) -> None:
    path = tmp_path / "sorting_configuration.xlsx"

    create_configuration_template(path)

    workbook = load_workbook(path)

    assert workbook.sheetnames == [
        "Instructions",
        "Pathologists",
        "CaseTypes",
        "Prefixes",
        "Hospitals",
        "RoutingOverrides",
        "AssignmentSettings",
        "Lists",
    ]
    assert workbook["Lists"].sheet_state == "hidden"

    assert [
        cell.value
        for cell in workbook["Pathologists"][1]
    ] == [
        "pathologist_id",
        "display_name",
        "subspecialties",
    ]

    assert [
        cell.value
        for cell in workbook["CaseTypes"][1]
    ] == [
        "case_type",
        "subspecialty",
        "requirement",
    ]

    assert workbook["Pathologists"].freeze_panes == "A2"
    assert workbook["CaseTypes"].freeze_panes == "A2"
    assert [
        cell.value
        for cell in workbook["AssignmentSettings"][1]
    ] == [
        "met_weight_per_pathologist",
        "wh_starting_weight",
    ]
    assert [
        cell.value
        for cell in workbook["AssignmentSettings"][2]
    ] == [200, 400]

    workbook.close()


def test_daily_template_has_expected_sheets(
    tmp_path: Path,
) -> None:
    path = tmp_path / "daily_sorting.xlsx"

    create_daily_template(path)

    workbook = load_workbook(path)

    assert workbook.sheetnames == [
        "Instructions",
        "Accessions",
        "Staffing",
        "Lists",
    ]
    assert workbook["Lists"].sheet_state == "hidden"

    assert [
        cell.value
        for cell in workbook["Accessions"][1]
    ] == [
        "accession_number",
        "prefix",
        "case_type",
        "hospital",
        "weight",
    ]

    staffing_headers = [
        cell.value
        for cell in workbook["Staffing"][1]
    ]
    assert staffing_headers == [
        "location",
        "pathologist_1",
        "pathologist_2",
        "pathologist_3",
        "pathologist_4",
        "pathologist_5",
        "pathologist_6",
    ]
    assert [
        workbook["Staffing"].cell(row=row, column=1).value
        for row in range(2, 8)
    ] == [
        "OLOL",
        "BRG",
        "WH",
        "MET",
        "TEXAS",
        "OMEGA",
    ]

    assert workbook["Accessions"].freeze_panes == "A2"
    assert workbook["Staffing"].freeze_panes == "A2"

    workbook.close()


def test_configuration_template_contains_validations(
    tmp_path: Path,
) -> None:
    path = tmp_path / "sorting_configuration.xlsx"

    create_configuration_template(path)

    workbook = load_workbook(path)

    case_type_validations = list(
        workbook["CaseTypes"].data_validations.dataValidation
    )
    prefix_validations = list(
        workbook["Prefixes"].data_validations.dataValidation
    )

    assert len(case_type_validations) >= 2
    assert len(prefix_validations) >= 2

    workbook.close()


def test_daily_template_contains_validations(
    tmp_path: Path,
) -> None:
    path = tmp_path / "daily_sorting.xlsx"

    create_daily_template(path)

    workbook = load_workbook(path)

    accession_validations = list(
        workbook["Accessions"].data_validations.dataValidation
    )
    staffing_validations = list(
        workbook["Staffing"].data_validations.dataValidation
    )

    assert len(accession_validations) == 3
    assert len(staffing_validations) == 1

    workbook.close()


def test_blank_templates_are_accepted_by_loader(
    tmp_path: Path,
) -> None:
    configuration_path, daily_path = create_sorting_templates(
        tmp_path
    )

    data = load_sorting_workbooks(
        configuration_path=configuration_path,
        daily_path=daily_path,
    )

    assert data.configuration.pathologists == ()
    assert data.configuration.case_type_rules == ()
    assert data.configuration.prefix_rules == ()
    assert data.configuration.hospital_rules == ()
    assert (
        data.configuration.assignment_settings.met_weight_per_pathologist
        == Decimal("200")
    )
    assert (
        data.configuration.assignment_settings.wh_starting_weight
        == Decimal("400")
    )
    assert data.daily.accessions == ()
    assert data.daily.staffing == ()

def test_routing_override_template_includes_weight_cap(tmp_path: Path) -> None:
    path = tmp_path / "sorting_configuration.xlsx"
    create_configuration_template(path)
    workbook = load_workbook(path)
    worksheet = workbook["RoutingOverrides"]

    assert [cell.value for cell in worksheet[1]][-1] == "weight_cap"
    assert (
        "preferred_until_weight_cap"
        in {
            workbook["Lists"].cell(row=row, column=3).value
            for row in range(2, workbook["Lists"].max_row + 1)
        }
    )

def test_daily_template_uses_configuration_pathologist_dropdowns(
    tmp_path: Path,
) -> None:
    configuration_path = tmp_path / "sorting_configuration.xlsx"
    daily_path = tmp_path / "daily_sorting.xlsx"

    create_configuration_template(configuration_path)
    workbook = load_workbook(configuration_path)
    worksheet = workbook["Pathologists"]
    worksheet["A2"] = "JS"
    worksheet["B2"] = "Dr. Smith"
    worksheet["C2"] = "GI"
    worksheet.append(["AB", "Dr. Brown", "BREAST"])
    workbook.save(configuration_path)
    workbook.close()

    create_daily_template(
        daily_path,
        configuration_path=configuration_path,
    )

    workbook = load_workbook(daily_path)
    lists = workbook["Lists"]
    staffing = workbook["Staffing"]

    assert lists["B2"].value is None
    assert lists["B3"].value == "JS"
    assert lists["B4"].value == "AB"
    assert "PathologistIDs" in workbook.defined_names

    staffing_validations = list(
        staffing.data_validations.dataValidation
    )
    assert len(staffing_validations) == 1
    assert staffing_validations[0].formula1 == "=PathologistIDs"

    workbook.close()

