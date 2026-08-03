"""Tests for Excel sorting-result reports."""

from decimal import Decimal
from pathlib import Path
from types import MappingProxyType

from openpyxl import load_workbook  # type: ignore[import-untyped]

from pgl_sorting_engine import (
    Accession,
    AssignmentMethod,
    AssignmentResult,
    EligibilityResult,
    LocationName,
    LocationSortingSummary,
    RoutingOverrideMode,
    RoutingOverrideRule,
    SortingRunResult,
    SubspecialtyRequirement,
    UnassignedAccession,
    create_sorting_report,
)


def _eligibility(accession: Accession, location: LocationName) -> EligibilityResult:
    return EligibilityResult(
        accession=accession,
        eligible_locations=frozenset({location}),
        preferred_locations=(),
        required_location=None,
        subspecialty=None,
        subspecialty_requirement=SubspecialtyRequirement.NOT_REQUIRED,
        exclusion_reasons=MappingProxyType({}),
        decision_notes=("Eligible for test assignment.",),
    )


def _assignment(
    accession_number: str,
    prefix: str,
    case_type: str,
    location: LocationName,
    weight: str = "1",
) -> AssignmentResult:
    accession = Accession(
        accession_number=accession_number,
        prefix=prefix,
        case_type=case_type,
        hospital="Hospital A",
        weight=Decimal(weight),
    )
    return AssignmentResult(
        accession=accession,
        location=location,
        method=AssignmentMethod.WEIGHT_BALANCED,
        eligibility=_eligibility(accession, location),
        assigned_weight_before=Decimal("0"),
        assigned_weight_after=Decimal(weight),
        target_weight=Decimal("10"),
        decision_notes=(f"Selected {location.value}.",),
    )


def _result() -> SortingRunResult:
    assignments = (
        _assignment("S26-1", "AB", "GI", LocationName.OLOL, "1.5"),
        _assignment("S26-2", "PG", "GI", LocationName.OLOL, "2.25"),
        _assignment("S26-3", "AB", "BR", LocationName.BRG),
        _assignment("S26-4", "PG", "BR", LocationName.WH),
        _assignment("S26-5", "AB", "GI", LocationName.MET),
    )
    summaries = {
        location: LocationSortingSummary(
            location=location,
            number_of_pathologists=1,
            target_weight=(
                None
                if location in {LocationName.TEXAS, LocationName.OMEGA}
                else Decimal("10")
            ),
            accession_count=sum(
                assignment.location is location for assignment in assignments
            ),
            assigned_weight=sum(
                (
                    assignment.accession.weight
                    for assignment in assignments
                    if assignment.location is location
                ),
                start=Decimal("0"),
            ),
            starting_weight=(
                Decimal("400") if location is LocationName.WH else Decimal("0")
            ),
        )
        for location in LocationName
    }
    unassigned_accession = Accession(
        accession_number="S26-U1",
        prefix="AB",
        case_type="GI",
        hospital="Hospital A",
        weight=Decimal("2"),
    )
    return SortingRunResult(
        input_accession_count=6,
        assignments=assignments,
        unassigned_accessions=(
            UnassignedAccession(
                accession=unassigned_accession,
                error_code="TEST_ERROR",
                summary="Test unassigned case.",
                details=("Manual review required.",),
            ),
        ),
        location_summaries=MappingProxyType(summaries),
    )


def test_create_sorting_report_writes_expected_sheets(tmp_path: Path) -> None:
    output_path = tmp_path / "sorting_results.xlsx"

    returned_path = create_sorting_report(_result(), output_path)

    assert returned_path == output_path
    workbook = load_workbook(output_path, data_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "Assignments",
        "Unassigned",
        "Audit",
        "Routing Override Matches",
        "Distribution Grids",
    ]
    assert workbook["Assignments"]["A2"].value == "S26-1"
    assert workbook["Unassigned"]["F2"].value == "TEST_ERROR"
    assert workbook["Summary"]["B3"].value == 5


def test_distribution_grids_sum_weight_and_do_not_freeze(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "sorting_results.xlsx"
    create_sorting_report(_result(), output_path)
    workbook = load_workbook(output_path, data_only=True)
    worksheet = workbook["Distribution Grids"]

    total_title_cell = next(
        cell
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value == "TOTAL: Weight by Case Type and Prefix"
    )
    olol_title_cell = next(
        cell
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value == "OLOL: Weight by Case Type and Prefix"
    )

    assert total_title_cell.row < olol_title_cell.row

    total_header_row = total_title_cell.row + 1
    total_headers = {
        worksheet.cell(
            row=total_header_row,
            column=column,
        ).value: column
        for column in range(1, worksheet.max_column + 1)
    }
    total_gi_row = next(
        row
        for row in range(
            total_header_row + 1,
            worksheet.max_row + 1,
        )
        if worksheet.cell(row=row, column=1).value == "GI"
    )

    assert (
        worksheet.cell(
            row=total_gi_row,
            column=total_headers["AB"],
        ).value
        == 2.5
    )
    assert (
        worksheet.cell(
            row=total_gi_row,
            column=total_headers["PG"],
        ).value
        == 2.25
    )
    assert (
        worksheet.cell(
            row=total_gi_row,
            column=total_headers["Total"],
        ).value
        == 4.75
    )

    header_row = olol_title_cell.row + 1
    headers = {
        worksheet.cell(row=header_row, column=column).value: column
        for column in range(1, worksheet.max_column + 1)
    }
    gi_row = next(
        row
        for row in range(header_row + 1, worksheet.max_row + 1)
        if worksheet.cell(row=row, column=1).value == "GI"
    )

    assert (
        worksheet.cell(
            row=gi_row,
            column=headers["AB"],
        ).value
        == 1.5
    )

    assert (
        worksheet.cell(
            row=gi_row,
            column=headers["PG"],
        ).value
        == 2.25
    )

    assert (
        worksheet.cell(
            row=gi_row,
            column=headers["Total"],
        ).value
        == 3.75
    )

    assert worksheet.freeze_panes is None


def test_report_writes_routing_override_match_sheet(tmp_path: Path) -> None:
    accession = Accession(
        accession_number="S26-R1",
        prefix="PG",
        case_type="GI",
        hospital="Lane",
        weight=Decimal("3"),
    )
    rule = RoutingOverrideRule(
        rule_name="Lane PG-GI to MET",
        hospital="Lane",
        prefix="PG",
        case_type="GI",
        mode=RoutingOverrideMode.PREFERRED_UNTIL_TARGET,
        destination_location=LocationName.MET,
    )
    eligibility = EligibilityResult(
        accession=accession,
        eligible_locations=frozenset({LocationName.MET, LocationName.OLOL}),
        preferred_locations=(),
        required_location=None,
        subspecialty="GI",
        subspecialty_requirement=SubspecialtyRequirement.PREFERRED,
        exclusion_reasons=MappingProxyType({}),
        decision_notes=("Matched override.",),
        override_rule=rule,
        override_activated=True,
        preferred_until_target_location=LocationName.MET,
        override_notes=("MET was below target.",),
    )
    assignment = AssignmentResult(
        accession=accession,
        location=LocationName.MET,
        method=AssignmentMethod.WEIGHT_BALANCED,
        eligibility=eligibility,
        assigned_weight_before=Decimal("8"),
        assigned_weight_after=Decimal("11"),
        target_weight=Decimal("10"),
        decision_notes=("Final case crossed target.",),
        override_applied=True,
        override_application_notes=("Override applied.",),
    )
    summaries = {
        location: LocationSortingSummary(
            location=location,
            number_of_pathologists=1,
            target_weight=(
                None
                if location in {LocationName.TEXAS, LocationName.OMEGA}
                else Decimal("10")
            ),
            accession_count=(1 if location is LocationName.MET else 0),
            assigned_weight=(
                Decimal("3") if location is LocationName.MET else Decimal("0")
            ),
        )
        for location in LocationName
    }
    result = SortingRunResult(
        input_accession_count=1,
        assignments=(assignment,),
        unassigned_accessions=(),
        location_summaries=MappingProxyType(summaries),
    )
    output_path = tmp_path / "routing_report.xlsx"
    create_sorting_report(result, output_path)

    workbook = load_workbook(output_path, data_only=True)
    worksheet = workbook["Routing Override Matches"]
    assert worksheet["A2"].value == "S26-R1"
    assert worksheet["F2"].value == "Lane PG-GI to MET"
    assert worksheet["I2"].value == "Yes"
    assert worksheet["M2"].value == "MET"


def test_report_writes_weight_cap_audit_values(tmp_path: Path) -> None:
    accession = Accession(
        accession_number="S26-CAP",
        prefix="PG",
        case_type="GI",
        hospital="Lane",
        weight=Decimal("15"),
    )
    rule = RoutingOverrideRule(
        rule_name="PG-GI to MET up to 40",
        prefix="PG",
        case_type="GI",
        mode=RoutingOverrideMode.PREFERRED_UNTIL_WEIGHT_CAP,
        destination_location=LocationName.MET,
        weight_cap=Decimal("40"),
    )
    eligibility = EligibilityResult(
        accession=accession,
        eligible_locations=frozenset({LocationName.MET, LocationName.OLOL}),
        preferred_locations=(),
        required_location=None,
        subspecialty=None,
        subspecialty_requirement=SubspecialtyRequirement.NOT_REQUIRED,
        exclusion_reasons=MappingProxyType({}),
        decision_notes=("Matched capped override.",),
        override_rule=rule,
        override_activated=True,
        preferred_until_weight_cap_location=LocationName.MET,
        override_notes=("Strict cap applied.",),
    )
    assignment = AssignmentResult(
        accession=accession,
        location=LocationName.MET,
        method=AssignmentMethod.WEIGHT_BALANCED,
        eligibility=eligibility,
        assigned_weight_before=Decimal("10"),
        assigned_weight_after=Decimal("25"),
        target_weight=Decimal("100"),
        decision_notes=("Capped preference selected MET.",),
        override_applied=True,
        override_application_notes=("Override applied.",),
        override_matching_weight_before=Decimal("25"),
        override_matching_weight_after=Decimal("40"),
        override_weight_cap=Decimal("40"),
    )
    summaries = {
        location: LocationSortingSummary(
            location=location,
            number_of_pathologists=1,
            target_weight=(
                None
                if location in {LocationName.TEXAS, LocationName.OMEGA}
                else Decimal("100")
            ),
            accession_count=(1 if location is LocationName.MET else 0),
            assigned_weight=(
                Decimal("15")
                if location is LocationName.MET
                else Decimal("0")
            ),
        )
        for location in LocationName
    }
    result = SortingRunResult(
        input_accession_count=1,
        assignments=(assignment,),
        unassigned_accessions=(),
        location_summaries=MappingProxyType(summaries),
    )
    output_path = tmp_path / "cap_report.xlsx"
    create_sorting_report(result, output_path)

    workbook = load_workbook(output_path, data_only=True)
    assignments = workbook["Assignments"]
    assignment_headers = {
        cell.value: cell.column for cell in assignments[1]
    }
    assert assignments.cell(
        row=2,
        column=assignment_headers["override_weight_cap"],
    ).value == 40
    assert assignments.cell(
        row=2,
        column=assignment_headers["override_matching_weight_before"],
    ).value == 25
    assert assignments.cell(
        row=2,
        column=assignment_headers["override_matching_weight_after"],
    ).value == 40

    matches = workbook["Routing Override Matches"]
    match_headers = {cell.value: cell.column for cell in matches[1]}
    assert matches.cell(
        row=2,
        column=match_headers["weight_cap"],
    ).value == 40
