"""Load configuration and daily sorting data from separate Excel workbooks."""

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from pgl_sorting_engine.assignment import SortingEngine
from pgl_sorting_engine.eligibility import EligibilityService
from pgl_sorting_engine.enums import (
    LocationName,
    SubspecialtyRequirement,
)
from pgl_sorting_engine.exceptions import (
    ConfigurationError,
    SortingEngineError,
    SpreadsheetIssue,
    SpreadsheetValidationError,
)
from pgl_sorting_engine.models import (
    Accession,
    DailyLocationStaffing,
    HospitalRoutingRule,
    Pathologist,
)
from pgl_sorting_engine.rules import (
    CaseTypeRule,
    PrefixRoutingRule,
    RoutingRuleSet,
)
from pgl_sorting_engine.staffing import DailySortingContext

CONFIGURATION_WORKBOOK = "CONFIGURATION"
DAILY_WORKBOOK = "DAILY"

PATHOLOGIST_HEADERS = (
    "pathologist_id",
    "display_name",
    "subspecialties",
)

CASE_TYPE_HEADERS = (
    "case_type",
    "subspecialty",
    "requirement",
)

PREFIX_HEADERS = (
    "prefix",
    "allowed_locations",
    "required_location",
    "preferred_locations",
)

HOSPITAL_HEADERS = (
    "hospital",
    "allowed_locations",
    "required_location",
)

ACCESSION_HEADERS = (
    "accession_number",
    "prefix",
    "case_type",
    "hospital",
    "weight",
)

STAFFING_HEADERS = (
    "location",
    "pathologist_id",
)


@dataclass(frozen=True, slots=True)
class SortingConfigurationData:
    """Stable pathologist and routing configuration."""

    pathologists: tuple[Pathologist, ...]
    case_type_rules: tuple[CaseTypeRule, ...]
    prefix_rules: tuple[PrefixRoutingRule, ...]
    hospital_rules: tuple[HospitalRoutingRule, ...]

    def build_rule_set(self) -> RoutingRuleSet:
        """Build the validated routing-rule collection."""
        return RoutingRuleSet(
            case_type_rules=self.case_type_rules,
            prefix_rules=self.prefix_rules,
            hospital_rules=self.hospital_rules,
        )


@dataclass(frozen=True, slots=True)
class DailySortingData:
    """Accessions and staffing information for one sorting day."""

    accessions: tuple[Accession, ...]
    staffing: tuple[DailyLocationStaffing, ...]


@dataclass(frozen=True, slots=True)
class SortingInputData:
    """Combined configuration and daily sorting input."""

    configuration: SortingConfigurationData
    daily: DailySortingData

    @property
    def accessions(self) -> tuple[Accession, ...]:
        """Return the daily accession collection."""
        return self.daily.accessions

    def build_staffing_context(self) -> DailySortingContext:
        """Build today's staffing and location capabilities."""
        return DailySortingContext(
            pathologists=self.configuration.pathologists,
            staffing=self.daily.staffing,
        )

    def build_engine(self) -> SortingEngine:
        """Build a complete sorting engine."""
        return SortingEngine(
            eligibility_service=EligibilityService(
                rules=self.configuration.build_rule_set(),
                staffing_context=self.build_staffing_context(),
            )
        )


def load_sorting_workbooks(
    configuration_path: str | Path,
    daily_path: str | Path,
) -> SortingInputData:
    """
    Load and validate separate configuration and daily workbooks.

    Args:
        configuration_path: Workbook containing stable routing configuration.
        daily_path: Workbook containing today's accessions and staffing.

    Returns:
        Combined, validated sorting input data.

    Raises:
        SpreadsheetValidationError: If either workbook contains invalid data.
    """
    issues: list[SpreadsheetIssue] = []

    configuration_workbook = _open_workbook(
        path=configuration_path,
        workbook_label=CONFIGURATION_WORKBOOK,
        issues=issues,
    )
    daily_workbook = _open_workbook(
        path=daily_path,
        workbook_label=DAILY_WORKBOOK,
        issues=issues,
    )

    if configuration_workbook is None or daily_workbook is None:
        if configuration_workbook is not None:
            configuration_workbook.close()

        if daily_workbook is not None:
            daily_workbook.close()

        raise SpreadsheetValidationError(tuple(issues))

    try:
        pathologists, pathologist_rows = _load_pathologists(
            configuration_workbook,
            issues,
        )
        case_type_rules = _load_case_type_rules(
            configuration_workbook,
            issues,
        )
        prefix_rules = _load_prefix_rules(
            configuration_workbook,
            issues,
        )
        hospital_rules = _load_hospital_rules(
            configuration_workbook,
            issues,
        )

        accessions, accession_rows = _load_accessions(
            daily_workbook,
            issues,
        )
        staffing = _load_staffing(
            daily_workbook,
            known_pathologist_ids=frozenset(pathologist_rows),
            issues=issues,
        )
    finally:
        configuration_workbook.close()
        daily_workbook.close()

    configuration = SortingConfigurationData(
        pathologists=tuple(pathologists),
        case_type_rules=tuple(case_type_rules),
        prefix_rules=tuple(prefix_rules),
        hospital_rules=tuple(hospital_rules),
    )

    daily = DailySortingData(
        accessions=tuple(accessions),
        staffing=tuple(staffing),
    )

    data = SortingInputData(
        configuration=configuration,
        daily=daily,
    )

    rule_set = _validate_configuration(
        configuration=configuration,
        issues=issues,
    )

    _validate_staffing(
        data=data,
        issues=issues,
    )

    if rule_set is not None:
        _validate_accession_references(
            accessions=data.accessions,
            accession_rows=accession_rows,
            rule_set=rule_set,
            issues=issues,
        )

    if issues:
        raise SpreadsheetValidationError(tuple(issues))

    return data


def _open_workbook(
    path: str | Path,
    workbook_label: str,
    issues: list[SpreadsheetIssue],
) -> Any | None:
    """Open one workbook and report a clear error if it fails."""
    workbook_path = Path(path).expanduser()

    if not workbook_path.exists():
        issues.append(
            SpreadsheetIssue(
                workbook=workbook_label,
                sheet=None,
                row_number=None,
                message=f"File does not exist: {workbook_path}.",
            )
        )
        return None

    try:
        return load_workbook(
            filename=workbook_path,
            read_only=True,
            data_only=True,
        )
    except (
        OSError,
        ValueError,
        BadZipFile,
        InvalidFileException,
    ) as exc:
        issues.append(
            SpreadsheetIssue(
                workbook=workbook_label,
                sheet=None,
                row_number=None,
                message=f"Could not open {workbook_path}: {exc}",
            )
        )
        return None


def _normalize_header(value: object) -> str:
    """Normalize an Excel column heading."""
    if value is None:
        return ""

    text = str(value).strip().lower().replace("-", " ")
    return "_".join(text.split())


def _is_blank(value: object) -> bool:
    """Return whether a spreadsheet cell is blank."""
    return value is None or (
        isinstance(value, str)
        and not value.strip()
    )


def _required_text(
    value: object,
    field_name: str,
) -> str:
    """Return required spreadsheet text."""
    if _is_blank(value):
        raise ValueError(f"{field_name} cannot be blank.")

    text = str(value).strip()

    if not text:
        raise ValueError(f"{field_name} cannot be blank.")

    return text


def _optional_text(value: object) -> str | None:
    """Return optional spreadsheet text."""
    if _is_blank(value):
        return None

    text = str(value).strip()
    return text or None


def _parse_positive_decimal(
    value: object,
    field_name: str,
) -> Decimal:
    """Parse a positive, finite Decimal value."""
    if _is_blank(value):
        raise ValueError(f"{field_name} cannot be blank.")

    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(
            f"{field_name} must be numeric; received {value!r}."
        ) from exc

    if not parsed.is_finite():
        raise ValueError(f"{field_name} must be finite.")

    if parsed <= 0:
        raise ValueError(f"{field_name} must be greater than zero.")

    return parsed


def _split_list(value: object) -> tuple[str, ...]:
    """
    Split comma- or semicolon-separated spreadsheet values.

    Examples:
        OLOL; BRG; MET
        GI, LIVER
    """
    text = _optional_text(value)

    if text is None:
        return ()

    normalized = text.replace(";", ",")

    return tuple(
        item.strip()
        for item in normalized.split(",")
        if item.strip()
    )


def _parse_location(value: object) -> LocationName:
    """Parse one location value."""
    text = _required_text(value, "Location").upper()

    try:
        return LocationName(text)
    except ValueError as exc:
        allowed_values = ", ".join(
            location.value
            for location in LocationName
        )

        raise ValueError(
            f"Unknown location {text!r}. "
            f"Expected one of: {allowed_values}."
        ) from exc


def _parse_optional_location(
    value: object,
) -> LocationName | None:
    """Parse an optional location value."""
    text = _optional_text(value)

    if text is None:
        return None

    return _parse_location(text)


def _parse_location_set(
    value: object,
) -> frozenset[LocationName]:
    """Parse a nonempty set of locations."""
    items = _split_list(value)

    if not items:
        raise ValueError(
            "Allowed locations must contain at least one location."
        )

    return frozenset(
        _parse_location(item)
        for item in items
    )


def _parse_location_tuple(
    value: object,
) -> tuple[LocationName, ...]:
    """Parse an ordered list of preferred locations."""
    return tuple(
        _parse_location(item)
        for item in _split_list(value)
    )


def _parse_requirement(
    value: object,
) -> SubspecialtyRequirement:
    """Parse a case-type subspecialty requirement."""
    text = _required_text(
        value,
        "Subspecialty requirement",
    )

    normalized = (
        text.lower()
        .replace("-", "_")
        .replace(" ", "_")
    )

    try:
        return SubspecialtyRequirement(normalized)
    except ValueError as exc:
        allowed_values = ", ".join(
            requirement.value
            for requirement in SubspecialtyRequirement
        )

        raise ValueError(
            f"Unknown requirement {text!r}. "
            f"Expected one of: {allowed_values}."
        ) from exc


def _sheet_rows(
    workbook: Any,
    workbook_label: str,
    sheet_name: str,
    required_headers: tuple[str, ...],
    issues: list[SpreadsheetIssue],
) -> list[tuple[int, dict[str, object]]]:
    """Read nonblank rows using normalized column headings."""
    if sheet_name not in workbook.sheetnames:
        issues.append(
            SpreadsheetIssue(
                workbook=workbook_label,
                sheet=sheet_name,
                row_number=None,
                message="Required worksheet is missing.",
            )
        )
        return []

    worksheet = workbook[sheet_name]
    row_iterator = worksheet.iter_rows(values_only=True)
    header_row = next(row_iterator, None)

    if header_row is None:
        issues.append(
            SpreadsheetIssue(
                workbook=workbook_label,
                sheet=sheet_name,
                row_number=None,
                message="Worksheet is empty.",
            )
        )
        return []

    header_indexes: dict[str, int] = {}

    for column_index, header_value in enumerate(header_row):
        normalized_header = _normalize_header(header_value)

        if not normalized_header:
            continue

        if normalized_header in header_indexes:
            issues.append(
                SpreadsheetIssue(
                    workbook=workbook_label,
                    sheet=sheet_name,
                    row_number=1,
                    message=(
                        f"Duplicate column heading "
                        f"{normalized_header!r}."
                    ),
                )
            )
            continue

        header_indexes[normalized_header] = column_index

    missing_headers = [
        header
        for header in required_headers
        if header not in header_indexes
    ]

    if missing_headers:
        issues.append(
            SpreadsheetIssue(
                workbook=workbook_label,
                sheet=sheet_name,
                row_number=1,
                message=(
                    "Missing required columns: "
                    + ", ".join(missing_headers)
                    + "."
                ),
            )
        )
        return []

    results: list[tuple[int, dict[str, object]]] = []

    for row_number, row in enumerate(
        row_iterator,
        start=2,
    ):
        if all(_is_blank(value) for value in row):
            continue

        values: dict[str, object] = {}

        for header in required_headers:
            column_index = header_indexes[header]

            if column_index >= len(row):
                values[header] = None
            else:
                values[header] = row[column_index]

        results.append((row_number, values))

    return results


def _append_row_issue(
    issues: list[SpreadsheetIssue],
    workbook: str,
    sheet: str,
    row_number: int,
    exc: Exception,
) -> None:
    """Record a row-specific spreadsheet issue."""
    issues.append(
        SpreadsheetIssue(
            workbook=workbook,
            sheet=sheet,
            row_number=row_number,
            message=str(exc),
        )
    )


def _load_pathologists(
    workbook: Any,
    issues: list[SpreadsheetIssue],
) -> tuple[list[Pathologist], dict[str, int]]:
    """Load the stable pathologist roster."""
    records: list[Pathologist] = []
    seen_ids: dict[str, int] = {}

    for row_number, values in _sheet_rows(
        workbook=workbook,
        workbook_label=CONFIGURATION_WORKBOOK,
        sheet_name="Pathologists",
        required_headers=PATHOLOGIST_HEADERS,
        issues=issues,
    ):
        try:
            pathologist = Pathologist(
                pathologist_id=_required_text(
                    values["pathologist_id"],
                    "Pathologist ID",
                ),
                display_name=_required_text(
                    values["display_name"],
                    "Display name",
                ),
                subspecialties=frozenset(
                    _split_list(values["subspecialties"])
                ),
            )

            previous_row = seen_ids.get(
                pathologist.pathologist_id
            )

            if previous_row is not None:
                raise ValueError(
                    f"Duplicate pathologist ID "
                    f"{pathologist.pathologist_id!r}; "
                    f"first seen on row {previous_row}."
                )

            seen_ids[pathologist.pathologist_id] = row_number
            records.append(pathologist)

        except (TypeError, ValueError, ConfigurationError) as exc:
            _append_row_issue(
                issues=issues,
                workbook=CONFIGURATION_WORKBOOK,
                sheet="Pathologists",
                row_number=row_number,
                exc=exc,
            )

    return records, seen_ids


def _load_case_type_rules(
    workbook: Any,
    issues: list[SpreadsheetIssue],
) -> list[CaseTypeRule]:
    """Load case-type subspecialty rules."""
    records: list[CaseTypeRule] = []
    seen_codes: dict[str, int] = {}

    for row_number, values in _sheet_rows(
        workbook=workbook,
        workbook_label=CONFIGURATION_WORKBOOK,
        sheet_name="CaseTypes",
        required_headers=CASE_TYPE_HEADERS,
        issues=issues,
    ):
        try:
            rule = CaseTypeRule(
                case_type=_required_text(
                    values["case_type"],
                    "Case type",
                ),
                subspecialty=_optional_text(
                    values["subspecialty"]
                ),
                requirement=_parse_requirement(
                    values["requirement"]
                ),
            )

            previous_row = seen_codes.get(rule.case_type)

            if previous_row is not None:
                raise ValueError(
                    f"Duplicate case type {rule.case_type!r}; "
                    f"first seen on row {previous_row}."
                )

            seen_codes[rule.case_type] = row_number
            records.append(rule)

        except (TypeError, ValueError, ConfigurationError) as exc:
            _append_row_issue(
                issues=issues,
                workbook=CONFIGURATION_WORKBOOK,
                sheet="CaseTypes",
                row_number=row_number,
                exc=exc,
            )

    return records


def _load_prefix_rules(
    workbook: Any,
    issues: list[SpreadsheetIssue],
) -> list[PrefixRoutingRule]:
    """Load prefix routing rules."""
    records: list[PrefixRoutingRule] = []
    seen_codes: dict[str, int] = {}

    for row_number, values in _sheet_rows(
        workbook=workbook,
        workbook_label=CONFIGURATION_WORKBOOK,
        sheet_name="Prefixes",
        required_headers=PREFIX_HEADERS,
        issues=issues,
    ):
        try:
            rule = PrefixRoutingRule(
                prefix=_required_text(
                    values["prefix"],
                    "Prefix",
                ),
                allowed_locations=_parse_location_set(
                    values["allowed_locations"]
                ),
                required_location=_parse_optional_location(
                    values["required_location"]
                ),
                preferred_locations=_parse_location_tuple(
                    values["preferred_locations"]
                ),
            )

            previous_row = seen_codes.get(rule.prefix)

            if previous_row is not None:
                raise ValueError(
                    f"Duplicate prefix {rule.prefix!r}; "
                    f"first seen on row {previous_row}."
                )

            seen_codes[rule.prefix] = row_number
            records.append(rule)

        except (TypeError, ValueError, ConfigurationError) as exc:
            _append_row_issue(
                issues=issues,
                workbook=CONFIGURATION_WORKBOOK,
                sheet="Prefixes",
                row_number=row_number,
                exc=exc,
            )

    return records


def _load_hospital_rules(
    workbook: Any,
    issues: list[SpreadsheetIssue],
) -> list[HospitalRoutingRule]:
    """Load hospital-origin routing rules."""
    records: list[HospitalRoutingRule] = []
    seen_hospitals: dict[str, int] = {}

    for row_number, values in _sheet_rows(
        workbook=workbook,
        workbook_label=CONFIGURATION_WORKBOOK,
        sheet_name="Hospitals",
        required_headers=HOSPITAL_HEADERS,
        issues=issues,
    ):
        try:
            rule = HospitalRoutingRule(
                hospital=_required_text(
                    values["hospital"],
                    "Hospital",
                ),
                allowed_locations=_parse_location_set(
                    values["allowed_locations"]
                ),
                required_location=_parse_optional_location(
                    values["required_location"]
                ),
            )

            previous_row = seen_hospitals.get(rule.hospital)

            if previous_row is not None:
                raise ValueError(
                    f"Duplicate hospital {rule.hospital!r}; "
                    f"first seen on row {previous_row}."
                )

            seen_hospitals[rule.hospital] = row_number
            records.append(rule)

        except (TypeError, ValueError, ConfigurationError) as exc:
            _append_row_issue(
                issues=issues,
                workbook=CONFIGURATION_WORKBOOK,
                sheet="Hospitals",
                row_number=row_number,
                exc=exc,
            )

    return records


def _load_accessions(
    workbook: Any,
    issues: list[SpreadsheetIssue],
) -> tuple[list[Accession], dict[str, int]]:
    """Load accessions from the daily workbook."""
    records: list[Accession] = []
    seen_accessions: dict[str, int] = {}

    for row_number, values in _sheet_rows(
        workbook=workbook,
        workbook_label=DAILY_WORKBOOK,
        sheet_name="Accessions",
        required_headers=ACCESSION_HEADERS,
        issues=issues,
    ):
        try:
            accession = Accession(
                accession_number=_required_text(
                    values["accession_number"],
                    "Accession number",
                ),
                prefix=_required_text(
                    values["prefix"],
                    "Prefix",
                ),
                case_type=_required_text(
                    values["case_type"],
                    "Case type",
                ),
                hospital=_required_text(
                    values["hospital"],
                    "Hospital",
                ),
                weight=_parse_positive_decimal(
                    values["weight"],
                    "Weight",
                ),
            )

            previous_row = seen_accessions.get(
                accession.accession_number
            )

            if previous_row is not None:
                raise ValueError(
                    f"Duplicate accession number "
                    f"{accession.accession_number!r}; "
                    f"first seen on row {previous_row}."
                )

            seen_accessions[
                accession.accession_number
            ] = row_number
            records.append(accession)

        except (TypeError, ValueError, ConfigurationError) as exc:
            _append_row_issue(
                issues=issues,
                workbook=DAILY_WORKBOOK,
                sheet="Accessions",
                row_number=row_number,
                exc=exc,
            )

    return records, seen_accessions


def _load_staffing(
    workbook: Any,
    known_pathologist_ids: frozenset[str],
    issues: list[SpreadsheetIssue],
) -> list[DailyLocationStaffing]:
    """Load one daily pathologist-location assignment per row."""
    assignments: dict[LocationName, list[str]] = {}
    seen_pathologists: dict[str, tuple[LocationName, int]] = {}

    for row_number, values in _sheet_rows(
        workbook=workbook,
        workbook_label=DAILY_WORKBOOK,
        sheet_name="Staffing",
        required_headers=STAFFING_HEADERS,
        issues=issues,
    ):
        try:
            location = _parse_location(values["location"])
            pathologist_id = _required_text(
                values["pathologist_id"],
                "Pathologist ID",
            ).upper()

            if pathologist_id not in known_pathologist_ids:
                raise ValueError(
                    f"Pathologist ID {pathologist_id!r} is not "
                    "present in the configuration workbook."
                )

            previous_assignment = seen_pathologists.get(
                pathologist_id
            )

            if previous_assignment is not None:
                previous_location, previous_row = previous_assignment

                raise ValueError(
                    f"Pathologist {pathologist_id!r} is already "
                    f"assigned to {previous_location.value} on "
                    f"row {previous_row}."
                )

            seen_pathologists[pathologist_id] = (
                location,
                row_number,
            )

            assignments.setdefault(location, []).append(
                pathologist_id
            )

        except (TypeError, ValueError, ConfigurationError) as exc:
            _append_row_issue(
                issues=issues,
                workbook=DAILY_WORKBOOK,
                sheet="Staffing",
                row_number=row_number,
                exc=exc,
            )

    records: list[DailyLocationStaffing] = []

    for location in LocationName:
        pathologist_ids = assignments.get(location)

        if pathologist_ids is None:
            continue

        try:
            records.append(
                DailyLocationStaffing(
                    location=location,
                    pathologist_ids=tuple(pathologist_ids),
                )
            )
        except (TypeError, ValueError, ConfigurationError) as exc:
            issues.append(
                SpreadsheetIssue(
                    workbook=DAILY_WORKBOOK,
                    sheet="Staffing",
                    row_number=None,
                    message=str(exc),
                )
            )

    return records


def _validate_configuration(
    configuration: SortingConfigurationData,
    issues: list[SpreadsheetIssue],
) -> RoutingRuleSet | None:
    """Validate configuration relationships and build the rule set."""
    try:
        return configuration.build_rule_set()
    except ConfigurationError as exc:
        issues.append(
            SpreadsheetIssue(
                workbook=CONFIGURATION_WORKBOOK,
                sheet=None,
                row_number=None,
                message=str(exc),
            )
        )
        return None


def _validate_staffing(
    data: SortingInputData,
    issues: list[SpreadsheetIssue],
) -> None:
    """Validate daily staffing against the configured roster."""
    try:
        data.build_staffing_context()
    except ConfigurationError as exc:
        issues.append(
            SpreadsheetIssue(
                workbook=DAILY_WORKBOOK,
                sheet="Staffing",
                row_number=None,
                message=str(exc),
            )
        )


def _validate_accession_references(
    accessions: tuple[Accession, ...],
    accession_rows: dict[str, int],
    rule_set: RoutingRuleSet,
    issues: list[SpreadsheetIssue],
) -> None:
    """Verify that each accession references configured routing values."""
    for accession in accessions:
        row_number = accession_rows[accession.accession_number]

        checks = (
            (
                rule_set.get_case_type_rule,
                accession.case_type,
            ),
            (
                rule_set.get_prefix_rule,
                accession.prefix,
            ),
            (
                rule_set.get_hospital_rule,
                accession.hospital,
            ),
        )

        for lookup, value in checks:
            try:
                lookup(value)
            except SortingEngineError as exc:
                issues.append(
                    SpreadsheetIssue(
                        workbook=DAILY_WORKBOOK,
                        sheet="Accessions",
                        row_number=row_number,
                        message=str(exc),
                    )
                )