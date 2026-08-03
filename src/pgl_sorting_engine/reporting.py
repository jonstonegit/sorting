"""Create Excel reports from completed sorting runs."""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Iterable, Sequence
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from pgl_sorting_engine.assignment import (
    AssignmentResult,
    SortingRunResult,
)
from pgl_sorting_engine.enums import LocationName

GRID_LOCATIONS = (
    LocationName.OLOL,
    LocationName.BRG,
    LocationName.WH,
    LocationName.MET,
)

ZERO_WEIGHT = Decimal("0")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)

SUMMARY_HEADERS = (
    "location",
    "pathologists",
    "cases",
    "assigned_weight",
    "starting_weight",
    "effective_weight",
    "target_weight",
    "variance_from_target",
    "assigned_weight_per_pathologist",
    "effective_weight_per_pathologist",
)

ASSIGNMENT_HEADERS = (
    "accession_number",
    "prefix",
    "case_type",
    "hospital",
    "weight",
    "assigned_location",
    "method",
    "target_weight",
    "override_rule",
    "override_mode",
    "override_applied",
    "override_destination",
    "override_notes",
    "decision_notes",
)

UNASSIGNED_HEADERS = (
    "accession_number",
    "prefix",
    "case_type",
    "hospital",
    "weight",
    "error_code",
    "summary",
    "details",
)

AUDIT_HEADERS = (
    "accession_number",
    "assigned_location",
    "method",
    "eligible_locations",
    "preferred_locations",
    "required_location",
    "subspecialty",
    "subspecialty_requirement",
    "assigned_weight_before",
    "assigned_weight_after",
    "target_weight",
    "override_rule",
    "override_mode",
    "override_matched",
    "override_activated",
    "override_applied",
    "override_destination",
    "override_notes",
    "eligibility_notes",
    "assignment_notes",
    "exclusion_reasons",
)


OVERRIDE_MATCH_HEADERS = (
    "accession_number",
    "prefix",
    "case_type",
    "hospital",
    "weight",
    "rule_name",
    "routing_mode",
    "override_activated",
    "override_applied",
    "destination_location",
    "preferred_locations",
    "required_subspecialty",
    "assigned_location",
    "assignment_method",
    "override_notes",
)


def create_sorting_report(
    result: SortingRunResult,
    output_path: str | Path,
) -> Path:
    """Write one completed sorting run to a formatted Excel workbook."""
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    _write_summary_sheet(summary_sheet, result)
    _write_assignments_sheet(
        workbook.create_sheet("Assignments"),
        result.assignments,
    )
    _write_unassigned_sheet(
        workbook.create_sheet("Unassigned"),
        result,
    )
    _write_audit_sheet(
        workbook.create_sheet("Audit"),
        result.assignments,
    )
    _write_override_matches_sheet(
        workbook.create_sheet("Routing Override Matches"),
        result.assignments,
    )
    _write_distribution_grids(
        workbook.create_sheet("Distribution Grids"),
        result.assignments,
    )

    workbook.save(destination)
    return destination


def _write_summary_sheet(
    worksheet: Worksheet,
    result: SortingRunResult,
) -> None:
    """Write overall run statistics and location summaries."""
    worksheet.append(["PGL Sorting Results"])
    worksheet["A1"].font = Font(size=16, bold=True)

    worksheet.append(["Input accessions", result.input_accession_count])
    worksheet.append(["Assigned accessions", result.assigned_accession_count])
    worksheet.append(["Unassigned accessions", result.unassigned_accession_count])
    worksheet.append(["Total assigned weight", _number(result.total_assigned_weight)])
    worksheet.append(["Total unassigned weight", _number(result.total_unassigned_weight)])

    header_row = 8
    _write_headers(worksheet, header_row, SUMMARY_HEADERS)

    for row_number, location in enumerate(LocationName, start=header_row + 1):
        summary = result.summary_for(location)
        worksheet.append(
            [
                location.value,
                summary.number_of_pathologists,
                summary.accession_count,
                _number(summary.assigned_weight),
                _number(summary.starting_weight),
                _number(summary.effective_weight),
                _optional_number(summary.target_weight),
                _optional_number(summary.variance_from_target),
                _optional_number(summary.assigned_weight_per_pathologist),
                _optional_number(summary.weight_per_pathologist),
            ]
        )
        _apply_number_formats(
            worksheet,
            row_number,
            columns=range(4, len(SUMMARY_HEADERS) + 1),
        )

    worksheet.freeze_panes = "A9"
    worksheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(SUMMARY_HEADERS))}"
        f"{header_row + len(LocationName)}"
    )
    _set_column_widths(
        worksheet,
        {
            "A": 18,
            "B": 14,
            "C": 11,
            "D": 18,
            "E": 18,
            "F": 18,
            "G": 16,
            "H": 21,
            "I": 31,
            "J": 32,
        },
    )


def _write_assignments_sheet(
    worksheet: Worksheet,
    assignments: Sequence[AssignmentResult],
) -> None:
    """Write the final destination for every assigned accession."""
    _write_headers(worksheet, 1, ASSIGNMENT_HEADERS)

    for row_number, assignment in enumerate(assignments, start=2):
        accession = assignment.accession
        worksheet.append(
            [
                accession.accession_number,
                accession.prefix,
                accession.case_type,
                accession.hospital,
                _number(accession.weight),
                assignment.location.value,
                assignment.method.value,
                _optional_number(assignment.target_weight),
                _override_rule_name(assignment),
                _override_mode(assignment),
                "Yes" if assignment.override_applied else "No",
                _override_destination(assignment),
                _override_notes(assignment),
                " | ".join(assignment.decision_notes),
            ]
        )
        _apply_number_formats(worksheet, row_number, columns=(5, 8))

    _configure_data_sheet(
        worksheet,
        headers=ASSIGNMENT_HEADERS,
        widths=(
            20, 10, 12, 36, 12, 20, 20, 16,
            32, 38, 18, 22, 80, 80,
        ),
    )


def _write_unassigned_sheet(
    worksheet: Worksheet,
    result: SortingRunResult,
) -> None:
    """Write accessions that require manual review."""
    _write_headers(worksheet, 1, UNASSIGNED_HEADERS)

    for row_number, item in enumerate(result.unassigned_accessions, start=2):
        accession = item.accession
        worksheet.append(
            [
                accession.accession_number,
                accession.prefix,
                accession.case_type,
                accession.hospital,
                _number(accession.weight),
                item.error_code,
                item.summary,
                " | ".join(item.details),
            ]
        )
        _apply_number_formats(worksheet, row_number, columns=(5,))

    _configure_data_sheet(
        worksheet,
        headers=UNASSIGNED_HEADERS,
        widths=(20, 10, 12, 36, 12, 28, 70, 90),
    )


def _write_audit_sheet(
    worksheet: Worksheet,
    assignments: Sequence[AssignmentResult],
) -> None:
    """Write eligibility and assignment reasoning for assigned cases."""
    _write_headers(worksheet, 1, AUDIT_HEADERS)

    for row_number, assignment in enumerate(assignments, start=2):
        eligibility = assignment.eligibility
        exclusion_reasons = " | ".join(
            (
                f"{location.value}: {'; '.join(reasons)}"
                for location, reasons in sorted(
                    eligibility.exclusion_reasons.items(),
                    key=lambda item: item[0].value,
                )
            )
        )

        worksheet.append(
            [
                assignment.accession.accession_number,
                assignment.location.value,
                assignment.method.value,
                _join_locations(eligibility.eligible_locations),
                _join_locations(eligibility.preferred_locations),
                (
                    eligibility.required_location.value
                    if eligibility.required_location is not None
                    else ""
                ),
                eligibility.subspecialty or "",
                eligibility.subspecialty_requirement.value,
                _number(assignment.assigned_weight_before),
                _number(assignment.assigned_weight_after),
                _optional_number(assignment.target_weight),
                _override_rule_name(assignment),
                _override_mode(assignment),
                "Yes" if eligibility.matched_override else "No",
                "Yes" if eligibility.override_activated else "No",
                "Yes" if assignment.override_applied else "No",
                _override_destination(assignment),
                _override_notes(assignment),
                " | ".join(eligibility.decision_notes),
                " | ".join(assignment.decision_notes),
                exclusion_reasons,
            ]
        )
        _apply_number_formats(worksheet, row_number, columns=(9, 10, 11))

    _configure_data_sheet(
        worksheet,
        headers=AUDIT_HEADERS,
        widths=(
            20, 20, 20, 35, 35, 20, 20, 26, 24, 23, 16,
            32, 38, 18, 20, 18, 22, 90, 90, 90, 100,
        ),
    )


def _write_override_matches_sheet(
    worksheet: Worksheet,
    assignments: Sequence[AssignmentResult],
) -> None:
    """Write every assigned accession that matched an override rule."""
    _write_headers(worksheet, 1, OVERRIDE_MATCH_HEADERS)

    matched = (
        assignment
        for assignment in assignments
        if assignment.eligibility.override_rule is not None
    )
    for row_number, assignment in enumerate(matched, start=2):
        rule = assignment.eligibility.override_rule
        if rule is None:
            continue
        worksheet.append(
            [
                assignment.accession.accession_number,
                assignment.accession.prefix,
                assignment.accession.case_type,
                assignment.accession.hospital,
                _number(assignment.accession.weight),
                rule.rule_name,
                rule.mode.value,
                (
                    "Yes"
                    if assignment.eligibility.override_activated
                    else "No"
                ),
                "Yes" if assignment.override_applied else "No",
                (
                    rule.destination_location.value
                    if rule.destination_location is not None
                    else ""
                ),
                _join_locations(rule.preferred_locations),
                rule.required_subspecialty or "",
                assignment.location.value,
                assignment.method.value,
                _override_notes(assignment),
            ]
        )
        _apply_number_formats(worksheet, row_number, columns=(5,))

    _configure_data_sheet(
        worksheet,
        headers=OVERRIDE_MATCH_HEADERS,
        widths=(
            20, 10, 12, 36, 12, 32, 38, 20, 18, 22,
            34, 28, 20, 20, 90,
        ),
    )


def _override_rule_name(assignment: AssignmentResult) -> str:
    rule = assignment.eligibility.override_rule
    return rule.rule_name if rule is not None else ""


def _override_mode(assignment: AssignmentResult) -> str:
    rule = assignment.eligibility.override_rule
    return rule.mode.value if rule is not None else ""


def _override_destination(assignment: AssignmentResult) -> str:
    rule = assignment.eligibility.override_rule
    if rule is None or rule.destination_location is None:
        return ""
    return rule.destination_location.value


def _override_notes(assignment: AssignmentResult) -> str:
    notes = [*assignment.eligibility.override_notes]
    notes.extend(assignment.override_application_notes)
    return " | ".join(dict.fromkeys(notes))


def _write_distribution_grids(
    worksheet: Worksheet,
    assignments: Sequence[AssignmentResult],
) -> None:
    """Write prefix-by-case-type weight grids for core locations."""
    relevant_assignments = tuple(
        assignment
        for assignment in assignments
        if assignment.location in GRID_LOCATIONS
    )

    prefixes = sorted(
        {
            assignment.accession.prefix
            for assignment in relevant_assignments
        }
    )

    case_types = sorted(
        {
            assignment.accession.case_type
            for assignment in relevant_assignments
        }
    )

    worksheet["A1"] = "Assigned Weight by Case Type and Prefix"
    worksheet["A1"].font = Font(size=16, bold=True)

    worksheet["A2"] = (
        "Each cell is the total assigned weight for the location, "
        "case type, and prefix."
    )

    start_row = 4

    for location in GRID_LOCATIONS:
        start_row = _write_location_grid(
            worksheet=worksheet,
            assignments=relevant_assignments,
            location=location,
            prefixes=prefixes,
            case_types=case_types,
            start_row=start_row,
        )

    worksheet.freeze_panes = None

    worksheet.column_dimensions["A"].width = 18

    for column_number in range(2, len(prefixes) + 3):
        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = 12

def _write_location_grid(
    worksheet: Worksheet,
    assignments: Sequence[AssignmentResult],
    location: LocationName,
    prefixes: Sequence[str],
    case_types: Sequence[str],
    start_row: int,
) -> int:
    """Write one location's weight matrix and return the next start row."""
    last_column = len(prefixes) + 2

    title = (
        f"{location.value}: Weight by Case Type and Prefix"
    )

    worksheet.cell(
        row=start_row,
        column=1,
        value=title,
    )

    worksheet.cell(
        row=start_row,
        column=1,
    ).font = BOLD_FONT

    worksheet.cell(
        row=start_row,
        column=1,
    ).fill = SECTION_FILL

    worksheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=max(last_column, 2),
    )

    header_row = start_row + 1
    headers = ["case_type", *prefixes, "Total"]

    _write_headers(
        worksheet,
        header_row,
        headers,
    )

    weights: defaultdict[
        tuple[str, str],
        Decimal,
    ] = defaultdict(lambda: ZERO_WEIGHT)

    for assignment in assignments:
        if assignment.location is not location:
            continue

        key = (
            assignment.accession.case_type,
            assignment.accession.prefix,
        )

        weights[key] += assignment.accession.weight

    for row_offset, case_type in enumerate(
        case_types,
        start=1,
    ):
        row_number = header_row + row_offset

        worksheet.cell(
            row=row_number,
            column=1,
            value=case_type,
        )

        worksheet.cell(
            row=row_number,
            column=1,
        ).font = BOLD_FONT

        row_total = ZERO_WEIGHT

        for prefix_offset, prefix in enumerate(
            prefixes,
            start=2,
        ):
            value = weights[
                (
                    case_type,
                    prefix,
                )
            ]

            worksheet.cell(
                row=row_number,
                column=prefix_offset,
                value=_number(value),
            )

            row_total += value

        worksheet.cell(
            row=row_number,
            column=last_column,
            value=_number(row_total),
        )

        _style_grid_row(
            worksheet,
            row_number,
            last_column,
        )

        _apply_number_formats(
            worksheet,
            row_number,
            columns=range(
                2,
                last_column + 1,
            ),
        )

    total_row = header_row + len(case_types) + 1

    worksheet.cell(
        row=total_row,
        column=1,
        value="Total",
    )

    for prefix_offset, prefix in enumerate(
        prefixes,
        start=2,
    ):
        value = sum(
            (
                weights[
                    (
                        case_type,
                        prefix,
                    )
                ]
                for case_type in case_types
            ),
            start=ZERO_WEIGHT,
        )

        worksheet.cell(
            row=total_row,
            column=prefix_offset,
            value=_number(value),
        )

    overall_total = sum(
        weights.values(),
        start=ZERO_WEIGHT,
    )

    worksheet.cell(
        row=total_row,
        column=last_column,
        value=_number(overall_total),
    )

    for cell in worksheet[total_row][:last_column]:
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center"
        )

    _apply_number_formats(
        worksheet,
        total_row,
        columns=range(
            2,
            last_column + 1,
        ),
    )

    return total_row + 3

    
def _write_headers(
    worksheet: Worksheet,
    row_number: int,
    headers: Sequence[str],
) -> None:
    """Write and style one header row."""
    for column_number, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=row_number,
            column=column_number,
            value=header,
        )
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _configure_data_sheet(
    worksheet: Worksheet,
    headers: Sequence[str],
    widths: Sequence[int],
) -> None:
    """Apply common filter, wrapping, and sizing to a data worksheet."""
    worksheet.freeze_panes = "A2"

    if worksheet.max_row > 1:
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{worksheet.max_row}"
        )

    for column_number, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_number)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_grid_row(
    worksheet: Worksheet,
    row_number: int,
    last_column: int,
) -> None:
    """Apply borders and alignment to one grid row."""
    for column_number in range(1, last_column + 1):
        cell = worksheet.cell(row=row_number, column=column_number)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _set_column_widths(
    worksheet: Worksheet,
    widths: dict[str, int],
) -> None:
    """Apply explicit worksheet column widths."""
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width


def _apply_number_formats(
    worksheet: Worksheet,
    row_number: int,
    columns: Iterable[int],
) -> None:
    """Format numeric report values consistently."""
    for column_number in columns:
        cell = worksheet.cell(row=row_number, column=column_number)
        if isinstance(cell.value, int | float):
            cell.number_format = "0.00"


def _join_locations(locations: Iterable[LocationName]) -> str:
    """Return a stable, comma-separated location list."""
    return ", ".join(
        location.value
        for location in sorted(locations, key=lambda item: item.value)
    )


def _number(value: Decimal) -> float:
    """Convert a Decimal into an Excel-compatible numeric value."""
    return float(value)


def _optional_number(value: Decimal | None) -> float | str:
    """Return an Excel number or N/A for locations without targets."""
    if value is None:
        return "N/A"

    return _number(value)
