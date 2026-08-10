"""Generate Excel templates for the PGL Sorting Engine."""

from argparse import ArgumentParser
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.comments import Comment  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import (  # type: ignore[import-untyped]
    get_column_letter,
)
from openpyxl.workbook.defined_name import (  # type: ignore[import-untyped]
    DefinedName,
)
from openpyxl.worksheet.datavalidation import (  # type: ignore[import-untyped]
    DataValidation,
)
from openpyxl.worksheet.table import (  # type: ignore[import-untyped]
    Table,
    TableStyleInfo,
)

CONFIGURATION_FILENAME = "sorting_configuration.xlsx"
DAILY_FILENAME = "daily_sorting.xlsx"

MAX_INPUT_ROW = 1000
MIN_STAFFING_SLOTS = 6
PATHOLOGIST_LIST_NAME = "PathologistIDs"

ASSIGNMENT_SETTINGS_HEADERS = (
    "met_weight_per_pathologist",
    "wh_starting_weight",
)

ASSIGNMENT_SETTINGS_DEFAULTS = (
    200,
    400,
)

LOCATIONS = (
    "OLOL",
    "BRG",
    "WH",
    "MET",
    "TEXAS",
    "OMEGA",
)

REQUIREMENTS = (
    "required",
    "preferred",
    "not_required",
)

ROUTING_MODES = (
    "identify_only",
    "always_required",
    "required_if_subspecialist_present",
    "preferred",
    "preferred_until_target",
    "preferred_until_weight_cap",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

SUBHEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

NOTE_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)

TITLE_FONT = Font(
    size=18,
    bold=True,
    color="1F4E78",
)

SECTION_FONT = Font(
    size=12,
    bold=True,
    color="1F4E78",
)

THIN_GRAY_BORDER = Border(
    bottom=Side(
        style="thin",
        color="B7B7B7",
    )
)


def create_sorting_templates(
    output_directory: str | Path,
) -> tuple[Path, Path]:
    """
    Create configuration and daily sorting workbook templates.

    Args:
        output_directory: Directory where both workbooks will be saved.

    Returns:
        Paths to the configuration and daily workbooks.
    """
    output_path = Path(output_directory).expanduser()
    output_path.mkdir(parents=True, exist_ok=True)

    configuration_path = output_path / CONFIGURATION_FILENAME
    daily_path = output_path / DAILY_FILENAME

    create_configuration_template(configuration_path)
    create_daily_template(
        daily_path,
        configuration_path=configuration_path,
    )

    return configuration_path, daily_path


def create_configuration_template(
    output_path: str | Path,
) -> Path:
    """
    Create the stable sorting-configuration workbook.

    The workbook contains:

    * Pathologists
    * CaseTypes
    * Prefixes
    * Hospitals
    * RoutingOverrides
    * AssignmentSettings
    """
    path = Path(output_path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()

    instructions = workbook.active
    instructions.title = "Instructions"

    pathologists = workbook.create_sheet("Pathologists")
    case_types = workbook.create_sheet("CaseTypes")
    prefixes = workbook.create_sheet("Prefixes")
    hospitals = workbook.create_sheet("Hospitals")
    routing_overrides = workbook.create_sheet("RoutingOverrides")
    assignment_settings = workbook.create_sheet("AssignmentSettings")
    lists = workbook.create_sheet("Lists")

    _build_assignment_settings_sheet(assignment_settings)
    _build_configuration_instructions(instructions)
    _build_configuration_lists(lists)

    _configure_input_sheet(
        worksheet=pathologists,
        headers=(
            "pathologist_id",
            "display_name",
            "subspecialties",
        ),
        widths=(18, 28, 36),
        table_name="PathologistsTable",
        tab_color="5B9BD5",
    )

    _configure_input_sheet(
        worksheet=case_types,
        headers=(
            "case_type",
            "subspecialty",
            "requirement",
        ),
        widths=(16, 28, 20),
        table_name="CaseTypesTable",
        tab_color="70AD47",
    )

    _configure_input_sheet(
        worksheet=prefixes,
        headers=(
            "prefix",
            "allowed_locations",
            "required_location",
            "preferred_locations",
        ),
        widths=(14, 34, 22, 34),
        table_name="PrefixesTable",
        tab_color="ED7D31",
    )

    _configure_input_sheet(
        worksheet=hospitals,
        headers=(
            "hospital",
            "allowed_locations",
            "required_location",
        ),
        widths=(34, 34, 22),
        table_name="HospitalsTable",
        tab_color="A5A5A5",
    )

    _configure_input_sheet(
        worksheet=routing_overrides,
        headers=(
            "rule_name",
            "hospital",
            "prefix",
            "case_type",
            "routing_mode",
            "destination_location",
            "preferred_locations",
            "required_subspecialty",
            "weight_cap",
        ),
        widths=(32, 32, 12, 14, 38, 24, 34, 28, 16),
        table_name="RoutingOverridesTable",
        tab_color="8064A2",
    )

    _add_two_letter_validation(
        worksheet=case_types,
        cell_range=f"A2:A{MAX_INPUT_ROW}",
        field_name="case type",
    )

    _add_two_letter_validation(
        worksheet=prefixes,
        cell_range=f"A2:A{MAX_INPUT_ROW}",
        field_name="prefix",
    )

    _add_two_letter_validation(
        worksheet=routing_overrides,
        cell_range=f"C2:C{MAX_INPUT_ROW}",
        field_name="prefix",
    )

    _add_two_letter_validation(
        worksheet=routing_overrides,
        cell_range=f"D2:D{MAX_INPUT_ROW}",
        field_name="case type",
    )

    location_formula = f'"{",".join(LOCATIONS)}"'

    requirement_formula = f'"{",".join(REQUIREMENTS)}"'
    routing_mode_formula = f'"{",".join(ROUTING_MODES)}"'

    _add_list_validation(
        worksheet=case_types,
        cell_range=f"C2:C{MAX_INPUT_ROW}",
        formula=requirement_formula,
        prompt="Select required, preferred, or not_required.",
        error="Select a valid subspecialty requirement.",
    )

    _add_list_validation(
        worksheet=prefixes,
        cell_range=f"C2:C{MAX_INPUT_ROW}",
        formula=location_formula,
        prompt="Select a required location or leave blank.",
        error="Select a valid location.",
    )

    _add_list_validation(
        worksheet=hospitals,
        cell_range=f"C2:C{MAX_INPUT_ROW}",
        formula=location_formula,
        prompt="Select a required location or leave blank.",
        error="Select a valid location.",
    )

    _add_list_validation(
        worksheet=routing_overrides,
        cell_range=f"E2:E{MAX_INPUT_ROW}",
        formula=routing_mode_formula,
        prompt="Select how the routing override should behave.",
        error="Select a valid routing mode.",
    )

    _add_list_validation(
        worksheet=routing_overrides,
        cell_range=f"F2:F{MAX_INPUT_ROW}",
        formula=location_formula,
        prompt=(
            "Select the destination used by required, until-target, or "
            "until-weight-cap rules."
        ),
        error="Select a valid location.",
    )

    _add_positive_decimal_validation(
        worksheet=routing_overrides,
        cell_range=f"I2:I{MAX_INPUT_ROW}",
        field_name="Weight cap",
    )

    _add_configuration_comments(
        pathologists=pathologists,
        case_types=case_types,
        prefixes=prefixes,
        hospitals=hospitals,
        routing_overrides=routing_overrides,
    )

    lists.sheet_state = "hidden"

    workbook.save(path)
    return path


def create_daily_template(
    output_path: str | Path,
    configuration_path: str | Path | None = None,
) -> Path:
    """
    Create the workbook used for one morning's sorting run.

    When ``configuration_path`` is provided, pathologist IDs are read from
    its Pathologists sheet and used as the Staffing dropdown choices.

    The workbook contains:
    * Accessions
    * Staffing
    """
    path = Path(output_path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)

    pathologist_ids = (
        _read_pathologist_ids(configuration_path)
        if configuration_path is not None
        else ()
    )
    staffing_slot_count = max(
        MIN_STAFFING_SLOTS,
        len(pathologist_ids),
    )

    workbook = Workbook()

    instructions = workbook.active
    instructions.title = "Instructions"

    accessions = workbook.create_sheet("Accessions")
    staffing = workbook.create_sheet("Staffing")
    lists = workbook.create_sheet("Lists")

    _build_daily_instructions(instructions)
    pathologist_list_last_row = _build_daily_lists(
        lists,
        pathologist_ids=pathologist_ids,
    )
    _define_pathologist_list(
        workbook,
        last_row=pathologist_list_last_row,
    )

    _configure_input_sheet(
        worksheet=accessions,
        headers=(
            "accession_number",
            "prefix",
            "case_type",
            "hospital",
            "weight",
        ),
        widths=(24, 14, 14, 34, 14),
        table_name="AccessionsTable",
        tab_color="4472C4",
    )
    _configure_staffing_sheet(
        worksheet=staffing,
        slot_count=staffing_slot_count,
    )

    _add_two_letter_validation(
        worksheet=accessions,
        cell_range=f"B2:B{MAX_INPUT_ROW}",
        field_name="prefix",
    )
    _add_two_letter_validation(
        worksheet=accessions,
        cell_range=f"C2:C{MAX_INPUT_ROW}",
        field_name="case type",
    )

    _add_positive_decimal_validation(
        worksheet=accessions,
        cell_range=f"E2:E{MAX_INPUT_ROW}",
        field_name="Accession weight",
    )

    last_staffing_column = get_column_letter(
        staffing_slot_count + 1
    )
    _add_list_validation(
        worksheet=staffing,
        cell_range=(
            f"B2:{last_staffing_column}{len(LOCATIONS) + 1}"
        ),
        formula=f"={PATHOLOGIST_LIST_NAME}",
        prompt=(
            "Select a pathologist for this location, or leave the cell "
            "blank."
        ),
        error=(
            "Select a pathologist from the configured roster or leave "
            "the cell blank."
        ),
    )

    accessions["E2"].number_format = "0.00"
    accessions["A1"].comment = Comment(
        "Enter each accession number only once.",
        "PGL Sorting Engine",
    )
    accessions["B1"].comment = Comment(
        "Enter the configured two-letter prefix.",
        "PGL Sorting Engine",
    )
    accessions["C1"].comment = Comment(
        "Enter the configured two-letter case type.",
        "PGL Sorting Engine",
    )
    accessions["D1"].comment = Comment(
        "Hospital name must match the configuration workbook.",
        "PGL Sorting Engine",
    )
    accessions["E1"].comment = Comment(
        "Enter a positive numeric workload weight.",
        "PGL Sorting Engine",
    )
    staffing["A1"].comment = Comment(
        "Locations are fixed. Choose today's pathologists across each row.",
        "PGL Sorting Engine",
    )
    staffing["B1"].comment = Comment(
        "Choose a configured pathologist ID from the dropdown. Blank is "
        "allowed; MET and OMEGA may have no pathologist on a given day.",
        "PGL Sorting Engine",
    )
    lists.sheet_state = "hidden"

    workbook.save(path)
    return path


def _read_pathologist_ids(
    configuration_path: str | Path,
) -> tuple[str, ...]:
    """Read unique pathologist IDs from a configuration workbook."""
    path = Path(configuration_path).expanduser()

    if not path.exists():
        raise FileNotFoundError(
            f"Configuration workbook does not exist: {path}."
        )

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        if "Pathologists" not in workbook.sheetnames:
            raise ValueError(
                "Configuration workbook is missing the Pathologists sheet."
            )

        worksheet = workbook["Pathologists"]
        row_iterator = worksheet.iter_rows(values_only=True)
        headers = next(row_iterator, None)

        if headers is None:
            raise ValueError(
                "The Pathologists sheet is empty."
            )

        normalized_headers = [
            ""
            if value is None
            else "_".join(
                str(value).strip().lower().replace("-", " ").split()
            )
            for value in headers
        ]

        try:
            pathologist_column = normalized_headers.index(
                "pathologist_id"
            )
        except ValueError as exc:
            raise ValueError(
                "The Pathologists sheet is missing the pathologist_id "
                "column."
            ) from exc

        pathologist_ids: list[str] = []
        seen_ids: set[str] = set()

        for row in row_iterator:
            if pathologist_column >= len(row):
                continue

            raw_value = row[pathologist_column]

            if raw_value is None:
                continue

            pathologist_id = str(raw_value).strip().upper()

            if not pathologist_id:
                continue

            if pathologist_id in seen_ids:
                raise ValueError(
                    f"Duplicate pathologist ID {pathologist_id!r} in "
                    "the configuration workbook."
                )

            seen_ids.add(pathologist_id)
            pathologist_ids.append(pathologist_id)

        return tuple(pathologist_ids)
    finally:
        workbook.close()


def _configure_staffing_sheet(
    worksheet: Any,
    slot_count: int,
) -> None:
    """Build one fixed staffing row per sorting location."""
    headers = (
        "location",
        *(
            f"pathologist_{slot_number}"
            for slot_number in range(1, slot_count + 1)
        ),
    )

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.sheet_properties.tabColor = "70AD47"
    worksheet.append(list(headers))

    for location in LOCATIONS:
        worksheet.append(
            [location, *([None] * slot_count)]
        )

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = THIN_GRAY_BORDER

    worksheet.row_dimensions[1].height = 24
    worksheet.column_dimensions["A"].width = 18

    for column_number in range(2, slot_count + 2):
        column_letter = get_column_letter(column_number)
        worksheet.column_dimensions[column_letter].width = 18

    for row_number in range(2, len(LOCATIONS) + 2):
        worksheet.row_dimensions[row_number].height = 22
        worksheet.cell(
            row=row_number,
            column=1,
        ).font = Font(bold=True)

    last_column = get_column_letter(slot_count + 1)
    table = Table(
        displayName="StaffingTable",
        ref=f"A1:{last_column}{len(LOCATIONS) + 1}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _define_pathologist_list(
    workbook: Any,
    last_row: int,
) -> None:
    """Create the workbook-level name used by staffing dropdowns."""
    defined_name = DefinedName(
        PATHOLOGIST_LIST_NAME,
        attr_text=(
            f"'Lists'!$B$2:$B${last_row}"
        ),
    )
    workbook.defined_names.add(defined_name)


def _configure_input_sheet(
    worksheet: Any,
    headers: tuple[str, ...],
    widths: tuple[int, ...],
    table_name: str,
    tab_color: str,
) -> None:
    """Apply the common layout used by an editable input worksheet."""
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.sheet_properties.tabColor = tab_color

    worksheet.append(list(headers))
    worksheet.append([None] * len(headers))

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = THIN_GRAY_BORDER

    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 20

    for column_number, width in enumerate(widths, start=1):
        column_letter = get_column_letter(column_number)
        worksheet.column_dimensions[column_letter].width = width

    last_column = get_column_letter(len(headers))
    table_reference = f"A1:{last_column}2"

    table = Table(
        displayName=table_name,
        ref=table_reference,
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    worksheet.add_table(table)

def _build_assignment_settings_sheet(
    worksheet: Any,
) -> None:
    """Build the editable assignment-settings worksheet."""

    worksheet.append(list(ASSIGNMENT_SETTINGS_HEADERS))
    worksheet.append(list(ASSIGNMENT_SETTINGS_DEFAULTS))

    worksheet.freeze_panes = "A2"

    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 24


def _add_list_validation(
    worksheet: Any,
    cell_range: str,
    formula: str,
    prompt: str,
    error: str,
) -> None:
    """Add single-choice dropdown validation to a range."""
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showInputMessage=True,
        showErrorMessage=True,
        errorStyle="stop",
    )
    validation.promptTitle = "PGL Sorting Engine"
    validation.prompt = prompt
    validation.errorTitle = "Invalid value"
    validation.error = error

    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def _add_two_letter_validation(
    worksheet: Any,
    cell_range: str,
    field_name: str,
) -> None:
    """Require exactly two characters when a code is entered."""
    validation = DataValidation(
        type="textLength",
        operator="equal",
        formula1="2",
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
        errorStyle="stop",
    )
    validation.promptTitle = "Two-letter code"
    validation.prompt = (
        f"Enter the configured two-letter {field_name}."
    )
    validation.errorTitle = "Invalid code"
    validation.error = (
        f"The {field_name} must contain exactly two characters."
    )

    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def _add_positive_decimal_validation(
    worksheet: Any,
    cell_range: str,
    field_name: str = "Weight",
) -> None:
    """Require a positive numeric value."""
    validation = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
        errorStyle="stop",
    )
    validation.promptTitle = field_name
    validation.prompt = (
        f"Enter a numeric {field_name.lower()} greater than zero."
    )
    validation.errorTitle = f"Invalid {field_name.lower()}"
    validation.error = (
        f"{field_name} must be a number greater than zero."
    )

    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def _build_configuration_lists(worksheet: Any) -> None:
    """Populate validation choices for the configuration workbook."""
    worksheet["A1"] = "Locations"
    worksheet["B1"] = "Requirements"
    worksheet["C1"] = "RoutingModes"

    for row_number, location in enumerate(LOCATIONS, start=2):
        worksheet.cell(
            row=row_number,
            column=1,
            value=location,
        )

    for row_number, requirement in enumerate(
        REQUIREMENTS,
        start=2,
    ):
        worksheet.cell(
            row=row_number,
            column=2,
            value=requirement,
        )

    for row_number, mode in enumerate(ROUTING_MODES, start=2):
        worksheet.cell(
            row=row_number,
            column=3,
            value=mode,
        )


def _build_daily_lists(
    worksheet: Any,
    pathologist_ids: tuple[str, ...] = (),
) -> int:
    """Populate hidden validation choices for the daily workbook."""
    worksheet["A1"] = "Locations"
    worksheet["B1"] = PATHOLOGIST_LIST_NAME

    for row_number, location in enumerate(LOCATIONS, start=2):
        worksheet.cell(
            row=row_number,
            column=1,
            value=location,
        )

    # B2 is intentionally blank so an empty dropdown choice is available.
    for row_number, pathologist_id in enumerate(
        pathologist_ids,
        start=3,
    ):
        worksheet.cell(
            row=row_number,
            column=2,
            value=pathologist_id,
        )

    return max(2, len(pathologist_ids) + 2)


def _add_configuration_comments(
    pathologists: Any,
    case_types: Any,
    prefixes: Any,
    hospitals: Any,
    routing_overrides: Any,
) -> None:
    """Explain fields that accept delimited lists or special values."""
    pathologists["A1"].comment = Comment(
        "Use a short unique identifier for each pathologist.",
        "PGL Sorting Engine",
    )
    pathologists["C1"].comment = Comment(
        "Separate multiple subspecialties with semicolons. "
        "Example: GI; LIVER",
        "PGL Sorting Engine",
    )

    case_types["A1"].comment = Comment(
        "The two-letter case type received from the LIS.",
        "PGL Sorting Engine",
    )
    case_types["B1"].comment = Comment(
        "Leave blank only when requirement is not_required.",
        "PGL Sorting Engine",
    )
    case_types["C1"].comment = Comment(
        "required: specialty must be present. "
        "preferred: specialty is favored. "
        "not_required: specialty does not affect eligibility.",
        "PGL Sorting Engine",
    )

    prefixes["B1"].comment = Comment(
        "Separate multiple locations with semicolons. "
        "Example: OLOL; BRG; MET",
        "PGL Sorting Engine",
    )
    prefixes["C1"].comment = Comment(
        "Optional hard destination. It must also appear in "
        "allowed_locations.",
        "PGL Sorting Engine",
    )
    prefixes["D1"].comment = Comment(
        "Optional ordered preferences separated by semicolons.",
        "PGL Sorting Engine",
    )

    hospitals["B1"].comment = Comment(
        "Separate multiple locations with semicolons.",
        "PGL Sorting Engine",
    )
    hospitals["C1"].comment = Comment(
        "Optional hard destination. Omega Hospital may use OMEGA here.",
        "PGL Sorting Engine",
    )

    routing_overrides["A1"].comment = Comment(
        "A readable name that will appear in the audit report.",
        "PGL Sorting Engine",
    )
    routing_overrides["B1"].comment = Comment(
        "Optional. Leave blank to apply the rule to all hospitals. An exact "
        "hospital rule takes priority over a general pair rule.",
        "PGL Sorting Engine",
    )
    routing_overrides["E1"].comment = Comment(
        "Modes: identify_only, always_required, "
        "required_if_subspecialist_present, preferred, "
        "preferred_until_target, or preferred_until_weight_cap.",
        "PGL Sorting Engine",
    )
    routing_overrides["F1"].comment = Comment(
        "Required for always_required, required_if_subspecialist_present, "
        "preferred_until_target, and preferred_until_weight_cap.",
        "PGL Sorting Engine",
    )
    routing_overrides["G1"].comment = Comment(
        "Used only for preferred mode. Separate ordered locations with "
        "semicolons.",
        "PGL Sorting Engine",
    )
    routing_overrides["H1"].comment = Comment(
        "Optional for required_if_subspecialist_present. When blank, the "
        "CaseTypes subspecialty is used.",
        "PGL Sorting Engine",
    )
    routing_overrides["I1"].comment = Comment(
        "Required only for preferred_until_weight_cap. Matching cases are "
        "preferred to the destination only when the new cumulative matching "
        "weight will remain at or below this cap.",
        "PGL Sorting Engine",
    )


def _build_configuration_instructions(
    worksheet: Any,
) -> None:
    """Create the configuration-workbook instruction sheet."""
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 88

    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = "PGL Sorting Engine — Configuration Workbook"
    worksheet["A1"].font = TITLE_FONT
    worksheet["A1"].alignment = Alignment(
        vertical="center",
    )
    worksheet.row_dimensions[1].height = 30

    rows = (
        (
            "Purpose",
            "Stores stable pathologist and routing configuration. "
            "This file should not be replaced each morning.",
        ),
        (
            "Pathologists",
            "One row per pathologist. Separate multiple "
            "subspecialties with semicolons.",
        ),
        (
            "CaseTypes",
            "Maps each two-letter case type to its subspecialty "
            "requirement.",
        ),
        (
            "Prefixes",
            "Defines allowed, required, and preferred work locations "
            "for each two-letter prefix.",
        ),
        (
            "Hospitals",
            "Defines which work locations may receive cases from each "
            "originating hospital.",
        ),
        (
            "RoutingOverrides",
            "Identifies or routes prefix/case-type pairs, optionally for one "
            "hospital. Hospital-specific rules take priority over general "
            "pair rules.",
        ),
        (
            "AssignmentSettings",
            "Sets MET weight per pathologist and WH starting weight. "
            "TEXAS and OMEGA rules are fixed in the application.",
        ),
        (
            "Important",
            "Do not rename sheets or column headings. Hospital names, "
            "prefixes, case types, and pathologist IDs must match the "
            "values used in the daily workbook.",
        ),
        (
            "Multiple values",
            "Use semicolons between multiple values, for example: "
            "OLOL; BRG; MET",
        ),
    )

    _write_instruction_rows(
        worksheet=worksheet,
        starting_row=3,
        rows=rows,
    )

    worksheet["A12"] = "Allowed locations"
    worksheet["A12"].font = SECTION_FONT
    worksheet["B12"] = "; ".join(LOCATIONS)

    worksheet["A14"] = "Requirement values"
    worksheet["A14"].font = SECTION_FONT
    worksheet["B14"] = "; ".join(REQUIREMENTS)

    worksheet["A16"] = "Routing override modes"
    worksheet["A16"].font = SECTION_FONT
    worksheet["B16"] = "; ".join(ROUTING_MODES)


def _build_daily_instructions(
    worksheet: Any,
) -> None:
    """Create the daily-workbook instruction sheet."""
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 88

    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = "PGL Sorting Engine — Daily Workbook"
    worksheet["A1"].font = TITLE_FONT
    worksheet["A1"].alignment = Alignment(
        vertical="center",
    )
    worksheet.row_dimensions[1].height = 30

    rows = (
        (
            "Purpose",
            "Contains the accessions and pathologist staffing for one "
            "morning's sorting run.",
        ),
        (
            "Accessions",
            "Enter one row per accession. Accession numbers must be "
            "unique.",
        ),
        (
            "Staffing",
            "Each sorting location has one row. Choose today's pathologists "
            "from the dropdown cells across that row.",
        ),
        (
            "Pathologist IDs",
            "Dropdown choices come from the Pathologists sheet when the "
            "daily template is created from a configuration workbook. "
            "Unused cells may be left blank, including all MET or OMEGA "
            "slots when those locations are unstaffed.",
        ),
        (
            "Routing values",
            "Prefix, case type, and hospital must exactly match the "
            "configuration workbook.",
        ),
        (
            "Privacy",
            "Do not include patient names or other unnecessary protected "
            "health information.",
        ),
        (
            "Important",
            "Do not rename sheets or column headings.",
        ),
    )

    _write_instruction_rows(
        worksheet=worksheet,
        starting_row=3,
        rows=rows,
    )

    worksheet["A12"] = "Sorting locations"
    worksheet["A12"].font = SECTION_FONT
    worksheet["B12"] = "; ".join(LOCATIONS)


def _write_instruction_rows(
    worksheet: Any,
    starting_row: int,
    rows: tuple[tuple[str, str], ...],
) -> None:
    """Write consistently formatted instruction rows."""
    for row_number, (heading, description) in enumerate(
        rows,
        start=starting_row,
    ):
        heading_cell = worksheet.cell(
            row=row_number,
            column=1,
            value=heading,
        )
        description_cell = worksheet.cell(
            row=row_number,
            column=2,
            value=description,
        )

        heading_cell.font = SECTION_FONT
        heading_cell.fill = SUBHEADER_FILL
        heading_cell.alignment = Alignment(
            vertical="top",
        )

        description_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        heading_cell.border = THIN_GRAY_BORDER
        description_cell.border = THIN_GRAY_BORDER

        worksheet.row_dimensions[row_number].height = 38

    warning_row = starting_row + len(rows) - 1
    worksheet.cell(
        row=warning_row,
        column=1,
    ).fill = NOTE_FILL
    worksheet.cell(
        row=warning_row,
        column=2,
    ).fill = NOTE_FILL


def _build_argument_parser() -> ArgumentParser:
    """Create the command-line argument parser."""
    parser = ArgumentParser(
        description=(
            "Create blank Excel templates for the "
            "PGL Sorting Engine."
        )
    )
    parser.add_argument(
        "--output-dir",
        default="templates",
        help=(
            "Directory where sorting_configuration.xlsx and "
            "daily_sorting.xlsx will be created."
        ),
    )
    return parser


def main() -> None:
    """Generate both templates from the command line."""
    parser = _build_argument_parser()
    arguments = parser.parse_args()

    configuration_path, daily_path = create_sorting_templates(
        arguments.output_dir
    )

    print(f"Created: {configuration_path}")
    print(f"Created: {daily_path}")


if __name__ == "__main__":
    main()